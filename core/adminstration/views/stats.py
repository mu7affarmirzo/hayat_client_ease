from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, F, ExpressionWrapper, DecimalField
from django.http import HttpResponse
from datetime import datetime, timedelta
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from .models import Session, Therapist


@login_required
def therapist_statistics(request):
    """
    View to display therapist statistics for a specified date range
    """
    # Default date range (last 30 days)
    end_date = datetime.now().date()
    start_date = end_date - timedelta(days=30)
    date_range = f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"

    # Get date range from request if provided
    if 'date_range' in request.GET and request.GET['date_range']:
        date_range = request.GET['date_range']
        start_str, end_str = date_range.split(' - ')
        start_date = datetime.strptime(start_str, '%d/%m/%Y').date()
        end_date = datetime.strptime(end_str, '%d/%m/%Y').date()

    # Get all therapists
    therapists = Therapist.objects.all()
    
    # Process statistics for each therapist
    stats = []
    total_sessions = 0
    total_amount = 0
    total_payout = 0
    
    for therapist in therapists:
        # Get sessions for this therapist in the date range
        sessions = Session.objects.filter(
            therapist=therapist,
            created_at__date__gte=start_date,
            created_at__date__lte=end_date,
        )
        
        # Calculate statistics
        session_count = sessions.aggregate(count=Count('id'))['count'] or 0
        amount = sessions.aggregate(total=Sum('total_price'))['total'] or 0
        
        # Calculate payout based on therapist rate
        rate = therapist.rate if hasattr(therapist, 'rate') else 0
        payout = amount * (rate / 100)
        
        # Skip therapists with no sessions
        if session_count == 0:
            continue
            
        # Add to totals
        total_sessions += session_count
        total_amount += amount
        total_payout += payout
        
        # Add to stats list
        stats.append({
            'full_name': therapist.full_name,
            'rate': rate,
            'session_count': session_count,
            'total_amount': amount,
            'payout_amount': payout,
        })
    
    # Calculate clinic profit
    profit = total_amount - total_payout
    
    context = {
        'therapists': stats,
        'start_date': start_date.strftime('%d/%m/%Y'),
        'end_date': end_date.strftime('%d/%m/%Y'),
        'date_range': date_range,
        'total_sessions': total_sessions,
        'total_amount': total_amount,
        'total_payout': total_payout,
        'profit': profit,
    }
    
    return render(request, 'reception_registration/therapist_statistics.html', context)


@login_required
def export_therapist_statistics(request):
    """
    Export therapist statistics to Excel
    """
    # Default date range (last 30 days)
    end_date = datetime.now().date()
    start_date = end_date - timedelta(days=30)
    
    # Get date range from request if provided
    if 'date_range' in request.GET and request.GET['date_range']:
        date_range = request.GET['date_range']
        start_str, end_str = date_range.split(' - ')
        start_date = datetime.strptime(start_str, '%d/%m/%Y').date()
        end_date = datetime.strptime(end_str, '%d/%m/%Y').date()
    
    # Get all therapists
    therapists = Therapist.objects.all()
    
    # Process statistics for each therapist
    stats = []
    total_sessions = 0
    total_amount = 0
    total_payout = 0
    
    for therapist in therapists:
        # Get sessions for this therapist in the date range
        sessions = Session.objects.filter(
            therapist=therapist,
            created_at__date__gte=start_date,
            created_at__date__lte=end_date,
        )
        
        # Calculate statistics
        session_count = sessions.aggregate(count=Count('id'))['count'] or 0
        amount = sessions.aggregate(total=Sum('total_price'))['total'] or 0
        
        # Calculate payout based on therapist rate
        rate = therapist.rate if hasattr(therapist, 'rate') else 0
        payout = amount * (rate / 100)
        
        # Skip therapists with no sessions
        if session_count == 0:
            continue
            
        # Add to totals
        total_sessions += session_count
        total_amount += amount
        total_payout += payout
        
        # Add to stats list
        stats.append({
            'full_name': therapist.full_name,
            'rate': rate,
            'session_count': session_count,
            'total_amount': amount,
            'payout_amount': payout,
        })
    
    # Create Excel workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Статистика терапевтов"
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 5
    worksheet.column_dimensions['B'].width = 25
    worksheet.column_dimensions['C'].width = 15
    worksheet.column_dimensions['D'].width = 15
    worksheet.column_dimensions['E'].width = 20
    worksheet.column_dimensions['F'].width = 20
    
    # Add title
    title = f"Статистика терапевтов за период {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
    worksheet['A1'] = title
    worksheet.merge_cells('A1:F1')
    title_cell = worksheet['A1']
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Add headers
    headers = ["#", "Терапевт", "Ставка (%)", "Кол-во сеансов", "Сумма (сум)", "К выплате (сум)"]
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=3, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Add data
    for row_num, therapist in enumerate(stats, 4):
        worksheet.cell(row=row_num, column=1).value = row_num - 3
        worksheet.cell(row=row_num, column=2).value = therapist['full_name']
        worksheet.cell(row=row_num, column=3).value = therapist['rate']
        worksheet.cell(row=row_num, column=4).value = therapist['session_count']
        worksheet.cell(row=row_num, column=5).value = therapist['total_amount']
        worksheet.cell(row=row_num, column=6).value = therapist['payout_amount']
    
    # Add totals
    total_row = len(stats) + 4
    worksheet.cell(row=total_row, column=1).value = "Итого:"
    worksheet.merge_cells(f'A{total_row}:C{total_row}')
    worksheet.cell(row=total_row, column=4).value = total_sessions
    worksheet.cell(row=total_row, column=5).value = total_amount
    worksheet.cell(row=total_row, column=6).value = total_payout
    
    # Make totals bold
    for col in range(1, 7):
        cell = worksheet.cell(row=total_row, column=col)
        cell.font = Font(bold=True)
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=therapist_statistics_{start_date.strftime("%Y%m%d")}-{end_date.strftime("%Y%m%d")}.xlsx'
    
    # Save workbook to response
    workbook.save(response)
    
    return response